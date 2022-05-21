'''module that does something'''
import sys
from openpyxl import load_workbook


class Color:
    '''docstring'''
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

# print(color.BOLD + 'Hello World !' + color.END)

def main():

    '''Docstring that is missing'''
    with open(sys.argv[1].rstrip(".xlsx")+'_ERRORS.txt', 'w', encoding="utf8") as f:

        #change these to match the Perf sheet cells
        cell_toe = 'AF22'
        cell_heel = 'AF26'
        tally_tab = 0

        workbookperf = load_workbook(filename=sys.argv[1], read_only=True, data_only=True)
        # grab number of stages from the sheet tab
        numstages = int(workbookperf.sheetnames[1][0:2:])

        # print the header for both the terminal and the file SS
        print(Color.BOLD + '\n\nLoading Perforations ...\n'+ Color.END)
        print('Perforations', file=f)
        print('\n',sys.argv[1][2::],'\n',workbookperf.sheetnames[1], file=f)
        print(sys.argv[1][2::],'\n')
        print(workbookperf.sheetnames[1])
        # pick the perf sheet with stages
        sheet = workbookperf[workbookperf.sheetnames[1]]
        perfdepth = []  # list because it is mutable
        for i, row in enumerate(sheet.iter_rows(min_row=8,
                                    max_row=(numstages*2)+7,
                                    min_col=3,
                                    max_col=17,
                                    values_only=True)):
            # pick every other row (evens)
            if i % 2 == 0:
                for cell in row:
                    if isinstance(cell, float) or isinstance(cell, int):
                        perfdepth.append(round(cell))
                        print(repr(round(cell)).rjust(5), end = ' ')
                        print(repr(round(cell)).rjust(5), end = ' ', file=f)
                print()
                print(file=f)
        # grab deepest and shallowest perfs
        deepperf = perfdepth[0]
        shallowperf = perfdepth[len(perfdepth)-1]


        print('\nLoading Collars ...')
        print()
        workbookcollars = load_workbook(filename=sys.argv[2], read_only=True, data_only=True)

        print(sys.argv[2][2::].strip(), workbookcollars.sheetnames[tally_tab],'\n')
        print('\n', sys.argv[2][2::].strip(), workbookcollars.sheetnames[tally_tab],'\n', file=f)
        # select the collars sheet
        # from the casing tally
        sheetcollar = workbookcollars[workbookcollars.sheetnames[tally_tab]]
        collardepth = []
        collarcount = 0
        for k, row in enumerate(sheetcollar.iter_rows(min_row=31,
                                    max_row=500,
                                    min_col=7, #set to row
                                    max_col=7,
                                    values_only=True), start=1):
            for cell in row:
                if isinstance(cell, float) or isinstance(cell, int):
                    if cell > 0:
                        collarcount +=1
                        collardepth.append(round(cell))
                        print(repr(round(cell)).rjust(5), end=' ')
                        print(repr(round(cell)).rjust(5), end=' ', file=f)
                        if k % 15 == 0:
                            print()
                            print(file=f)
        print('\n\n', collarcount, 'Collars found.')
        print('\n\n', collarcount, 'Collars found.', file = f)
        print('\nList of conflicts: \n')
        print('Collar      Perf \n', end='')
        print('------     ------ \n', end='')
        print('\nList of conflicts: \n', file=f)
        print('Collar      Perf \n', end='',file=f)
        print('------     ------ \n', end='',file=f)

        conflict = []  # create a conflict list for some project later
        for collar in tuple(collardepth):

            for perf in tuple(perfdepth):
                if int(collar) == int(perf):
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is the same')
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is the same',file = f)
                    conflict.append(perf)
                elif int(collar)-1 == int(perf):
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is +1 above')
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is +1 above',file = f)
                    conflict.append(perf)
                elif int(collar)-2 == int(perf):
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is +2 above')
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is +2 above',file = f)
                    conflict.append(perf)
                elif int(collar)+1 == int(perf):
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is -1 below')
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is -1 below',file = f)
                    conflict.append(perf)
                elif int(collar)+2 == int(perf):
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is -2 below')
                    print(repr(collar).rjust(5), '    ', repr(perf).rjust(5),
                    ' is -2 below',file = f)
                    conflict.append(perf)
        if len(conflict) == 0:
            print()
            print(file = f)
            print('No conflicts detected.')
            print('No conflicts detected.', file = f)
            print()
            print(file = f)
        else:
            print()
            print(file = f)
            print(len(conflict),'conflicts detected.')
            print(len(conflict),'conflicts detected.', file = f)
            print()
            print(file = f)

        print('Loading Survey ...')
        print()
        print('Survey', file = f)
        print(file = f)
        print(sys.argv[3][2::].strip())
        print(sys.argv[3][2::].strip(), file = f)
        workbooksurvey = load_workbook(filename=sys.argv[3], read_only=True, data_only=True)
        sheetsurvey = workbooksurvey[workbooksurvey.sheetnames[0]]
        surveykb = 0
        surveyheel = 0
        surveytoe = 0
        print(sheetsurvey.cell(8,9).value[3:5],
            'ft KB depth from definitive survey header found')
        print(sheetsurvey.cell(8,9).value[3:5],
            'ft KB depth from definitive survey header found', file = f)
        for i, row in enumerate(sheetsurvey.iter_rows(min_row=24,
                                    max_row=300,
                                    min_col=0,
                                    max_col=2,
                                    values_only=True)):
            for cell in row:
                if repr(cell)[1:3] == "KB":
                    surveykb = int(round(row[1]))
                    print(round(row[1]), 'ft KB depth from definitive survey data found')
                    print(round(row[1]), 'ft KB depth from definitive survey data found', file = f)
                elif repr(cell)[1:19] == "Cross Setback Heel":
                    surveyheel = round(row[1])
                    print(surveyheel, 'Cross Setback Heel depth found')
                    print(surveyheel, 'Cross Setback Heel depth found', file = f)
                elif repr(cell)[1:18] == "Cross Setback Toe":
                    surveytoe = round(row[1])
                    print(surveytoe, 'Cross Setback Toe depth found')
                    print(surveytoe, 'Cross Setback Toe depth found', file = f)
                #need to capture not finding cross setback toe and cross setbak errors
                # print(repr(cell)[1:16])

            # print(i, row)

    # print deep/shallow and summary of good/bad
        sheetsetback = workbookperf[workbookperf.sheetnames[0]]
        surveykb = int((sheetsurvey.cell(8,9).value[3:5]))
        surveyheelgl = surveyheel - surveykb
        surveytoegl = surveytoe - surveykb
        print('\nSummary')
        print('\nSummary', file = f)
        print('\nDeepest perf        Shallowest perf')
        print('\nDeepest perf        Shallowest perf',file=f)  
        print(' ', deepperf,'             ', shallowperf)
        print(' ', deepperf,'             ', shallowperf,file=f)
        print('\nToe Set-back        Heel set-back')
        print('\nToe Set-back        Heel set-back',file=f)
        print(' ',round(sheetsetback[cell_toe].value),
            '             ', round(sheetsetback[cell_heel].value))
        print(' ',round(sheetsetback[cell_toe].value),
            '             ', round(sheetsetback[cell_heel].value),file=f)
        print('\nSurvey Toe SB (GL)   Survey Heel SB (GL)')
        print('\nSurvey Toe SB (GL)   Survey Heel SB (GL)',file=f) 
        print(' ', surveytoegl,'             ', surveyheelgl)
        print(' ', surveytoegl,'             ', surveyheelgl,file=f)
        print()
        print(file = f)
        # Future -  need error handling made for this next part
        # AF22 and AF26 cell values might change 
        if int(deepperf) < int(sheetsetback[cell_toe].value):
            print('Toe perf is within toe set-back line, Toe perfs are good.')
            print('Toe perf is within toe set-back line, Toe perfs are good.',file=f)
        if int(deepperf) >= int(sheetsetback[cell_toe].value):
            print('ERROR, Toe perf is deeper than toe set-back.')
            print('ERROR, Toe perf is deeper than toe set-back.',file=f)
        if int(shallowperf) > int(sheetsetback[cell_heel].value):
            print('Heel perf is within heel set-back line, Heel perfs are good.')
            print('Heel perf is within heel set-back line, Heel perfs are good.',file=f)
        if int(shallowperf) <= int(sheetsetback[cell_heel].value):
            print('ERROR, Heel perf is shallower than heel set-back.')
            print('ERROR, Heel perf is shallower than heel set-back.',file=f)
        if int(surveyheelgl) != int(round(sheetsetback[cell_heel].value)):
            print('Warning!, Survey heel setback ', surveyheelgl,
                ', does not equal perf sheet setback. ',
                int(round(sheetsetback[cell_heel].value)),' ',
                surveyheelgl-int(round(sheetsetback[cell_heel].value)),'ft difference.')
            print('Warning!, Survey heel setback ', surveyheelgl,
                ', does not equal perf sheet setback. ',
                int(round(sheetsetback[cell_heel].value)),' ',
                surveyheelgl-int(round(sheetsetback[cell_heel].value)),'ft difference.', file=f)
        if int(surveytoegl) != int(round(sheetsetback[cell_toe].value)):
            print('Warning!, Survey toe setback ',surveytoegl,
                ', does not equal perf sheet toe setback. ',
                int(round(sheetsetback[cell_toe].value)),' ',
                surveytoegl-int(sheetsetback[cell_toe].value),'ft difference.')
            print('Warning!, Survey toe setback ',surveytoegl,
                ', does not equal perf sheet toe setback. ',
                int(round(sheetsetback[cell_toe].value)),' ',
                surveytoegl-int(sheetsetback[cell_toe].value),'ft difference.', file = f)
        print('\n')


        # print(conflict)
if __name__ == "__main__":
    main()
