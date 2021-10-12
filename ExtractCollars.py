import sys
from openpyxl import load_workbook
workbookCollars = load_workbook(filename=sys.argv[1], data_only=True)
print('\n',workbookCollars.sheetnames[1],'\n')
# with open(sys.argv[1], 'r') as f: 

sheet = workbookCollars[workbookCollars.sheetnames[1]]
collarDepth = []
for row in sheet.iter_rows(min_row=12,
                            max_row=400,
                            min_col=21,
                            max_col=21,
                            values_only=True):
     
    # print(row)
    for cell in row:
        if isinstance(cell, float):
            collarDepth.append(round(cell))   
            print(round(cell))
        
print(tuple(collarDepth))