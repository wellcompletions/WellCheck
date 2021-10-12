import sys
from openpyxl import load_workbook
workbook = load_workbook(filename=sys.argv[1], data_only=True)
print('\n',workbook.sheetnames[1],'\n')
# with open(sys.argv[1], 'r') as f: 

sheet = workbook[workbook.sheetnames[1]]
perfDepth = []
for i,row in enumerate(sheet.iter_rows(min_row=8,
                            max_row=77,
                            min_col=3,
                            max_col=15,
                            values_only=True)):
    
    if i % 2 == 0:
        print(i)
        for cell in row:
            
            perfDepth.append(round(cell))   
            print(round(cell), end = ' ')
        print()

    