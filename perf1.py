from openpyxl import load_workbook
workbook = load_workbook(filename="Perfs.xlsx", data_only=True)
print('\n',workbook.sheetnames[1],'\n')
sheet = workbook['35 -Stage Perf Design']
perfDepth = []
for row in sheet.iter_rows(min_row=8,
                            max_row=42,
                            min_col=3,
                            max_col=15,
                            values_only=True):
                        
    
    for cell in row:
        perfDepth.append(round(cell))   
        print(round(cell), end = ' ')
    print()
# print('\n\nPerforations Depths: ','\n',tuple(perfDepth), end=' ')

# read file with collar values
with open('collars.txt', "r") as f:
	listItems = f.read().split('\t\t\n')
# print(tuple(listItems))  

print('\n\nList of conflicts: \n')
print('Collar      Perf \n', end='')
print('-----      ----- \n', end='')

conflict = []
for collar in tuple(listItems):
     #print(collar)
     for perf in tuple(perfDepth):
        if int(collar) == int(perf):
            print(collar, '    ', perf, ' is same')
            conflict.append(perf)
            # conflict.append(int(perf)+2)
        elif int(collar)+1 == int(perf):
            print(collar, '    ', perf, ' is +1 above')
            # print(perf, ' oh no +1')
            conflict.append(perf)
            # conflict.append(int(perf)+2)
        elif int(collar)+2 == int(perf):
            print(collar, '    ', perf, ' is +2 above')
            # print(perf, ' oh no +2')
            conflict.append(perf)
        elif int(collar)-1 == int(perf):
            print(collar, '    ', perf, ' is -1 below')
            # print(perf, ' oh no -1 ')
            conflict.append(perf)
        elif int(collar)-2 == int(perf):
            print(collar, '    ', perf, ' is -2 below') 
            # print(perf, ' oh no -2 ')
            conflict.append(perf)
        
print('\n\n\n')
# print(conflict)

