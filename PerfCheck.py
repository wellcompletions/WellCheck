from openpyxl import load_workbook
import sys 

if __name__ == "__main__":
    newfilename = sys.argv[1].rstrip(".xlsx")+'_ERRORS.txt'

    with open(newfilename, 'w') as f:
          
        workbookPerf = load_workbook(filename=sys.argv[1], read_only=True, data_only=True)
        numstages = int(workbookPerf.sheetnames[1][0:2:])
        print('\n\nLoading Perforations ...\n')
        print('Perforations', file=f)
        print('\n',sys.argv[1][2::],'\n',workbookPerf.sheetnames[1], file=f)
        
        print(sys.argv[1][2::],'\n')
        print(workbookPerf.sheetnames[1])
        
        sheet = workbookPerf[workbookPerf.sheetnames[1]]
        perfDepth = []
        for i, row in enumerate(sheet.iter_rows(min_row=8,
                                    max_row=(numstages*2)+7,
                                    min_col=3,
                                    max_col=15,
                                    values_only=True)):
            
            if i % 2 == 0:
                for cell in row:            
                    perfDepth.append(round(cell))   
                    print(round(cell), end = ' ')
                    print(round(cell), end = ' ', file=f)
                print()
                print(file=f)

        deepPerf = perfDepth[0]
        shallowPerf = perfDepth[len(perfDepth)-1]
        
        
        print('\nLoading Collars ...')
        workbookCollars = load_workbook(filename=sys.argv[2], read_only=True, data_only=True)
        print()
        print(sys.argv[2][2::].strip(), workbookCollars.sheetnames[1],'\n')
        print('\n', sys.argv[2][2::].strip(), workbookCollars.sheetnames[1],'\n', file=f)

        sheetCollar = workbookCollars[workbookCollars.sheetnames[1]]
        collarDepth = []
        k = 0

        for row in sheetCollar.iter_rows(min_row=12,
                                    max_row=400,
                                    min_col=21,
                                    max_col=21,
                                    values_only=True):
            
            # print(row)
            for cell in row:
                k = k +1
                if isinstance(cell, float):
                    collarDepth.append(round(cell))   
                    print(round(cell), end=' ')
                    print(round(cell), end=' ', file=f)
                    if k % 13 == 0:
                        print()
                        print(file=f)
                
        # print(tuple(collarDepth))

        print('\n\nList of conflicts: \n')
        print('Collar      Perf \n', end='')
        print('------     ------ \n', end='')
        print('\n\nList of conflicts: \n', file=f)
        print('Collar      Perf \n', end='',file=f)
        print('------     ------ \n', end='',file=f)

        conflict = []
        for collar in tuple(collarDepth):
            #print(collar)
            for perf in tuple(perfDepth):
                if int(collar) == int(perf):
                    print(collar, '    ', perf, ' is same')
                    print(collar, '    ', perf, ' is same',file=f)
                    conflict.append(perf)
                elif int(collar)+1 == int(perf):
                    print(collar, '    ', perf, ' is +1 above')
                    print(collar, '    ', perf, ' is +1 above',file=f)
                    conflict.append(perf)
                elif int(collar)+2 == int(perf):
                    print(collar, '    ', perf, ' is +2 above')
                    print(collar, '    ', perf, ' is +2 above',file=f)
                    conflict.append(perf)
                elif int(collar)-1 == int(perf):
                    print(collar, '    ', perf, ' is -1 below')
                    print(collar, '    ', perf, ' is -1 below',file=f)
                    conflict.append(perf)
                elif int(collar)-2 == int(perf):
                    print(collar, '    ', perf, ' is -2 below')
                    print(collar, '    ', perf, ' is -2 below',file=f) 
                    conflict.append(perf)
    
        sheetSetBack = workbookPerf[workbookPerf.sheetnames[0]]

        print('\nDeepest perf    Shallowest perf')    
        print('\nDeepest perf    Shallowest perf',file=f)          
        print(' ', deepPerf,'         ', shallowPerf) 
        print(' ', deepPerf,'         ', shallowPerf,file=f)   
        print('\nToe Set-back    Heel set-back') 
        print('\nToe Set-back    Heel set-back',file=f) 
        print(' ',round(sheetSetBack['AF22'].value), '         ', round(sheetSetBack['AF26'].value),'\n')  
        print(' ',round(sheetSetBack['AF22'].value), '         ', round(sheetSetBack['AF26'].value),'\n',file=f) 


    # Future -  need error handling made for this next part 

        if int(deepPerf) < int(sheetSetBack['AF22'].value):
            print('Toe perf is within toe set-back line, perfs are good.')
            print('Toe perf is within toe set-back line, perfs are good.',file=f) 
        if int(deepPerf) >= int(sheetSetBack['AF22'].value):
            print('ERROR, Toe perf is deeper than toe set-back.')
            print('ERROR, Toe perf is deeper than toe set-back.',file=f)
        if int(shallowPerf) > int(sheetSetBack['AF27'].value):
            print('Heel perf is within heel set-back line, perfs are good.')
            print('Heel perf is within heel set-back line, perfs are good.',file=f)
        if int(shallowPerf) <= int(sheetSetBack['AF27'].value):
            print('ERROR, Heel perf is shallower than heel set-back.')
            print('ERROR, Heel perf is shallower than heel set-back.',file=f)
        print('\n')
        # print(conflict)

