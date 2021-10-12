from openpyxl import load_workbook
import sys 
print('\n\nLoading Perforations\n\n')
workbookPerf = load_workbook(filename=sys.argv[1], data_only=True)
print('\n',workbookPerf.sheetnames[1],'\n')
sheet = workbookPerf['35 -Stage Perf Design']
perfDepth = []
for i,row in enumerate(sheet.iter_rows(min_row=8,
                            max_row=77,
                            min_col=3,
                            max_col=15,
                            values_only=True)):
    
    if i % 2 == 0:
        for cell in row:            
            perfDepth.append(round(cell))   
            print(round(cell), end = ' ')
        print()

print('\n\nLoading Collars\n\n')
workbookCollars = load_workbook(filename=sys.argv[2], data_only=True)
print('\n',workbookCollars.sheetnames[1],'\n')
# with open(sys.argv[1], 'r') as f: 

sheetCollar = workbookCollars[workbookCollars.sheetnames[1]]
collarDepth = []
for row in sheetCollar.iter_rows(min_row=12,
                            max_row=400,
                            min_col=21,
                            max_col=21,
                            values_only=True):
     
    # print(row)
    for cell in row:
        if isinstance(cell, float):
            collarDepth.append(round(cell))   
            print(round(cell))
        
print(tuple(collarDepth))








# # read file with collar values
# with open('collars.txt', "r") as f:
# 	listItems = f.read().split('\t\t\n')
# # print(tuple(listItems))  

print('\n\nList of conflicts: \n')
print('Collar      Perf \n', end='')
print('------     ------ \n', end='')

conflict = []
for collar in tuple(collarDepth):
     #print(collar)
     for perf in tuple(perfDepth):
        if int(collar) == int(perf):
            print(collar, '    ', perf, ' is same')
            conflict.append(perf)
            # conflict.append(int(perf)+2)
        elif int(collar)+1 == int(perf):
            print(collar, '    ', perf, ' is +1 above')
            # print(perf, ' oh no +1')
            conflict.append(perf)
            # conflict.append(int(perf)+2)
        elif int(collar)+2 == int(perf):
            print(collar, '    ', perf, ' is +2 above')
            # print(perf, ' oh no +2')
            conflict.append(perf)
        elif int(collar)-1 == int(perf):
            print(collar, '    ', perf, ' is -1 below')
            # print(perf, ' oh no -1 ')
            conflict.append(perf)
        elif int(collar)-2 == int(perf):
            print(collar, '    ', perf, ' is -2 below') 
            # print(perf, ' oh no -2 ')
            conflict.append(perf)
        
print('\n\n\n')
# print(conflict)

